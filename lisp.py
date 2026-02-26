"""
GitHub repo sourcer (LISP-only, repo-centric) + owner enrichment + location geocoding/normalization.

- Searches GitHub repos for BASE_QUERY + EXTRA_QUALIFIERS (date range)
- For each repo: collects repo fields
- For each owner: fetches profile fields (cached) and enriches with:
  owner_name, owner_email, owner_location (raw), blog/website, X, LinkedIn, extra links
- Geocodes/normalizes the owner's location (raw text) into:
  owner_location_norm, owner_city, owner_region, owner_country, owner_country_code, owner_lat, owner_lon
  + provider + status
- Writes clickable Excel (or CSV fallback)

Providers:
- Google Geocoding API (if GOOGLE_MAPS_API_KEY set, or GEO_PROVIDER=google)
- Nominatim (OSM) fallback by default (or GEO_PROVIDER=nominatim)

Requires:
- requests, pandas, openpyxl
- optional: geopy (only needed if using Nominatim)

Terminal (VS Code):
  pip install -r requirements.txt
Env:
  setx GITHUB_TOKEN "...."
  (optional) setx GOOGLE_MAPS_API_KEY "...."
"""

import os
import re
import json
import time
import random
import typing as t
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote_plus, urlparse

import requests
import pandas as pd

import gspread
from google.oauth2.service_account import Credentials

SPREADSHEET_ID = "1OVr2EigkJ5ZHceilXGn-Zl8xpGTGPVxp8jKIEJnOgmo"
SHEET_GID = 1382489855
SERVICE_ACCOUNT_FILE = "google_service_account.json"

def get_gspread_worksheet():
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=scopes
    )
    client = gspread.authorize(creds)
    sh = client.open_by_key(SPREADSHEET_ID)

    # Find worksheet by gid
    for ws in sh.worksheets():
        if ws.id == SHEET_GID:
            return ws

    raise ValueError(f"Worksheet with gid={SHEET_GID} not found. Check the gid in your URL.")


from openpyxl import load_workbook
from openpyxl.styles import Font
from requests.exceptions import ReadTimeout, ConnectionError, HTTPError

from dotenv import load_dotenv
load_dotenv()

import os
print("TOKEN present:", bool(os.getenv("GITHUB_TOKEN")))



# ---------------- Config ----------------
GITHUB_API = "https://api.github.com"

BASE_QUERY = "lisp"
EXTRA_QUALIFIERS = "created:2023-01-01..2026-12-31"

MAX_REPOS = 500
PER_PAGE = 50
TIMEOUT = 20

DEFAULT_XLSX = "github_repos_lisp_with_owner_details.xlsx"

# Gentle pacing to reduce abuse detection on /search endpoints
PAGE_SLEEP_RANGE = (0.2, 0.8)  # seconds (randomized)

# Geocoding
GEO_PROVIDER = (os.getenv("GEO_PROVIDER") or "").strip().lower()  # "google" or "nominatim" (optional)
GOOGLE_MAPS_API_KEY = (os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()

# Cache file for geocoding results (disk)
GEO_CACHE_FILE = Path(__file__).with_name("geocode_cache.json")
_GEO_CACHE: dict[str, dict] = {}

# Common non-geocodable locations
_BAD_LOCATIONS = {
    "", "remote", "worldwide", "earth", "somewhere", "internet", "everywhere", "global", "online",
    "anywhere", "planet earth", "the internet", "github", "home",
}

# ---------------- Token -----------------
# Prefer env var GITHUB_TOKEN. Optionally load from token_1.py if you keep secrets locally.
try:
    from token_1 import GITHUB_TOKEN_2
except Exception:
    GITHUB_TOKEN_2 = ""

ENV_TOKEN = (os.getenv("GITHUB_TOKEN") or "").strip()
FILE_TOKEN = (GITHUB_TOKEN_2 or "").strip()
TOKEN = ENV_TOKEN or FILE_TOKEN

SESSION = requests.Session()
SESSION.headers.update({
    "Accept": "application/vnd.github+json",
    "User-Agent": "github-sourcer-script/5.1",
})

# Use Bearer (most robust across token types)
if TOKEN:
    SESSION.headers.update({"Authorization": f"Bearer {TOKEN}"})
else:
    print("No GitHub token found (GITHUB_TOKEN env var or token_1.py:GITHUB_TOKEN_2). "
          "Running unauthenticated => very low rate limit.")

# 🔍 DEBUG (DOIT être hors du if/else)
auth = SESSION.headers.get("Authorization", "")
print("TOKEN present:", bool(TOKEN))
print("Auth scheme:", auth.split(" ")[0] if auth else "NONE")
print("Auth length:", len(auth))

r = SESSION.get("https://api.github.com/rate_limit", timeout=20)
print("rate_limit status:", r.status_code)
print("rate_limit body:", r.text[:200])


# Owner enrichment (cached in-memory)
_OWNER_CACHE: dict[str, dict] = {}


# ---------------- Diagnostics ----------------
def _print_auth_diagnostics() -> None:
    source = "env:GITHUB_TOKEN" if ENV_TOKEN else ("token_1.py:GITHUB_TOKEN_2" if FILE_TOKEN else "none")
    print(f"Token source: {source}")
    print(f"Authorization header present: {'Authorization' in SESSION.headers}")


def _print_rate_limit_snapshot() -> None:
    try:
        r = SESSION.get(f"{GITHUB_API}/rate_limit", timeout=TIMEOUT)
        if r.status_code == 200:
            data = r.json() or {}
            core = (data.get("resources") or {}).get("core") or {}
            search = (data.get("resources") or {}).get("search") or {}
            print(f"RateLimit core: remaining={core.get('remaining')} reset={core.get('reset')}")
            print(f"RateLimit search: remaining={search.get('remaining')} reset={search.get('reset')}")
        else:
            print(f"RateLimit check failed: HTTP {r.status_code}")
    except Exception as e:
        print("RateLimit check failed:", e)


# ---------------- Helpers ----------------
def safe_output_path(filename: str) -> Path:
    p = Path(__file__).with_name(filename)
    try:
        p.touch(exist_ok=True)
        return p
    except PermissionError:
        return p.with_name(p.stem + "_new" + p.suffix)


def normalize_url(url: str) -> str:
    if not url:
        return ""
    url = str(url).strip()
    if not url:
        return ""
    if not url.lower().startswith(("http://", "https://")):
        url = "https://" + url
    try:
        return url if urlparse(url).netloc else ""
    except Exception:
        return ""


def urls_from_text(text: str) -> list[str]:
    if not text:
        return []
    urls = re.findall(r"(https?://[^\s)]+)", text, flags=re.IGNORECASE)
    out: list[str] = []
    for u in urls:
        nu = normalize_url(u.rstrip(".,);]}>\"'"))
        if nu and nu not in out:
            out.append(nu)
    return out


def extract_first_linkedin(*fields: str) -> str:
    for field in fields:
        if not field:
            continue
        urls = urls_from_text(field)
        for u in urls:
            if "linkedin.com" in u.lower():
                return u
        s = field.strip()
        if "linkedin.com" in s.lower():
            idx = s.lower().find("linkedin.com")
            candidate = s[idx:].split()[0].strip().rstrip(".,);]}>\"'")
            return normalize_url(candidate)
    return ""


def created_year_in_range(created_at: str, start_year: int = 2023, end_year: int = 2026) -> bool:
    if not created_at:
        return False
    try:
        dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
        return start_year <= dt.year <= end_year
    except Exception:
        return False


def write_excel_with_fallback(df: pd.DataFrame, filename: str) -> Path:
    out = safe_output_path(filename)
    try:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

        wb = load_workbook(out)
        ws = wb.active
        headers = {c.value: i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}

        link_cols = {
            "repo_url": "Repo",
            "owner_url": "Owner",
            "owner_blog": "Website",
            "owner_x": "X",
            "owner_linkedin": "LinkedIn",
        }

        for r in range(2, ws.max_row + 1):
            for col, label in link_cols.items():
                idx = headers.get(col)
                if not idx:
                    continue
                cell = ws.cell(row=r, column=idx)
                val = (cell.value or "").strip()
                if not val:
                    continue
                url = val.split(";")[0].strip()
                if url.lower().startswith(("http://", "https://")):
                    cell.hyperlink = url
                    cell.value = label
                    cell.font = Font(color="0563C1", underline="single")

        wb.save(out)
        print(f"Excel written: {out.resolve()} ({len(df)} rows)")
        return out

    except Exception as e:
        print("Excel failed, falling back to CSV:", e)
        out_csv = out.with_suffix(".csv")
        df.to_csv(out_csv, index=False)
        print(f"CSV written: {out_csv.resolve()} ({len(df)} rows)")
        return out_csv


def get(url: str) -> requests.Response:
    """Resilient GET with retries on transient errors + rate limiting."""
    max_attempts = 6
    base_sleep = 2.0
    last_exc: Exception | None = None

    for attempt in range(1, max_attempts + 1):
        try:
            resp = SESSION.get(url, timeout=TIMEOUT)

            # If auth is wrong, do NOT retry forever
            if resp.status_code == 401:
                resp.raise_for_status()

            # Rate limit handling (403 + reset header, sometimes 429 too)
            if resp.status_code in (403, 429):
                reset = resp.headers.get("X-RateLimit-Reset")
                remaining = resp.headers.get("X-RateLimit-Remaining")
                if reset and (remaining == "0" or remaining is None):
                    wait = max(0, int(reset) - int(time.time())) + 2
                    print(f"Rate limit hit. Sleeping {wait}s...")
                    time.sleep(wait)
                    continue

            if resp.status_code in (502, 503, 504):
                sleep = base_sleep * (2 ** (attempt - 1)) + random.uniform(0, 1)
                print(f"Transient {resp.status_code}. Retry {attempt}/{max_attempts} in {sleep:.1f}s")
                time.sleep(sleep)
                continue

            resp.raise_for_status()
            return resp

        except (ReadTimeout, ConnectionError, HTTPError) as e:
            last_exc = e
            sleep = base_sleep * (2 ** (attempt - 1)) + random.uniform(0, 1)
            print(f"Error {e}. Retry {attempt}/{max_attempts} in {sleep:.1f}s")
            time.sleep(sleep)

    raise RuntimeError(f"GET failed after retries: {last_exc}")


def build_query() -> str:
    q = BASE_QUERY.strip()
    if EXTRA_QUALIFIERS.strip():
        q = f"{q} {EXTRA_QUALIFIERS.strip()}"
    return q


def search_repositories(query: str) -> t.Iterable[dict]:
    page = 1
    while True:
        qp = quote_plus(query)
        url = f"{GITHUB_API}/search/repositories?q={qp}&per_page={PER_PAGE}&page={page}"
        data = get(url).json()
        items = data.get("items") or []
        if not items:
            break

        for repo in items:
            yield repo

        page += 1
        time.sleep(random.uniform(*PAGE_SLEEP_RANGE))


def fetch_owner(login: str) -> dict:
    if not login:
        return {}
    if login in _OWNER_CACHE:
        return _OWNER_CACHE[login]
    data = get(f"{GITHUB_API}/users/{login}").json()
    _OWNER_CACHE[login] = data or {}
    return _OWNER_CACHE[login]


def owner_fields(owner_json: dict) -> dict:
    name = (owner_json.get("name") or "").strip()
    email = (owner_json.get("email") or "").strip()
    location = (owner_json.get("location") or "").strip()
    blog_raw = (owner_json.get("blog") or "").strip()
    blog = normalize_url(blog_raw)
    bio = (owner_json.get("bio") or "").strip()
    twitter_user = (owner_json.get("twitter_username") or "").strip()
    x_url = normalize_url(f"https://twitter.com/{twitter_user}") if twitter_user else ""
    linkedin = extract_first_linkedin(blog_raw, bio)
    extra = urls_from_text(bio)
    for known in [blog, linkedin, x_url]:
        if known and known in extra:
            extra.remove(known)
    extra_links = "; ".join(extra)

    return {
        "owner_name": name,
        "owner_email": email,
        "owner_location": location,  # raw
        "owner_blog": blog,
        "owner_x": x_url,
        "owner_linkedin": linkedin,
        "owner_extra_links": extra_links,
    }


# ---------------- Geocoding (cache + providers) ----------------
def load_geo_cache() -> None:
    global _GEO_CACHE
    if GEO_CACHE_FILE.exists():
        try:
            _GEO_CACHE = json.loads(GEO_CACHE_FILE.read_text(encoding="utf-8"))
            if not isinstance(_GEO_CACHE, dict):
                _GEO_CACHE = {}
        except Exception:
            _GEO_CACHE = {}


def save_geo_cache() -> None:
    try:
        GEO_CACHE_FILE.write_text(
            json.dumps(_GEO_CACHE, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    except Exception:
        pass


def _geo_empty(provider: str = "") -> dict:
    return {
        "owner_location_norm": "",
        "owner_city": "",
        "owner_region": "",
        "owner_country": "",
        "owner_country_code": "",
        "owner_lat": "",
        "owner_lon": "",
        "owner_geocode_provider": provider,
        "owner_geocode_status": "EMPTY",
    }


def _geo_no_match(provider: str) -> dict:
    d = _geo_empty(provider)
    d["owner_geocode_status"] = "NO_MATCH"
    return d


def geocode_google(raw: str) -> dict:
    provider = "google"
    raw = (raw or "").strip()
    key = raw.lower()

    if not raw or key in _BAD_LOCATIONS:
        d = _geo_empty(provider)
        d["owner_geocode_status"] = "SKIPPED"
        return d

    if not GOOGLE_MAPS_API_KEY:
        d = _geo_empty(provider)
        d["owner_geocode_status"] = "NO_API_KEY"
        return d

    cache_key = f"{provider}:{key}"
    if cache_key in _GEO_CACHE:
        return _GEO_CACHE[cache_key]

    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": raw, "key": GOOGLE_MAPS_API_KEY}

    out = _geo_no_match(provider)
    try:
        r = requests.get(url, params=params, timeout=TIMEOUT)
        data = r.json() if r.content else {}
        status = (data.get("status") or "").upper()

        if status != "OK":
            out["owner_geocode_status"] = status or "ERROR"
            _GEO_CACHE[cache_key] = out
            return out

        results = data.get("results") or []
        if not results:
            _GEO_CACHE[cache_key] = out
            return out

        top = results[0]
        formatted = top.get("formatted_address") or ""
        geom = (top.get("geometry") or {}).get("location") or {}
        lat = geom.get("lat")
        lon = geom.get("lng")

        comps = top.get("address_components") or []
        comp_map = {}
        for c in comps:
            types = c.get("types") or []
            for ty in types:
                comp_map.setdefault(ty, c)

        def _long(ty: str) -> str:
            return (comp_map.get(ty) or {}).get("long_name") or ""

        def _short(ty: str) -> str:
            return (comp_map.get(ty) or {}).get("short_name") or ""

        city = _long("locality") or _long("postal_town") or _long("administrative_area_level_3")
        region = _long("administrative_area_level_1")
        country = _long("country")
        country_code = _short("country")

        out = {
            "owner_location_norm": formatted,
            "owner_city": city,
            "owner_region": region,
            "owner_country": country,
            "owner_country_code": country_code,
            "owner_lat": "" if lat is None else str(lat),
            "owner_lon": "" if lon is None else str(lon),
            "owner_geocode_provider": provider,
            "owner_geocode_status": "OK",
        }

    except Exception:
        out["owner_geocode_status"] = "ERROR"

    _GEO_CACHE[cache_key] = out
    return out


def geocode_nominatim(raw: str) -> dict:
    provider = "nominatim"
    raw = (raw or "").strip()
    key = raw.lower()

    if not raw or key in _BAD_LOCATIONS:
        d = _geo_empty(provider)
        d["owner_geocode_status"] = "SKIPPED"
        return d

    cache_key = f"{provider}:{key}"
    if cache_key in _GEO_CACHE:
        return _GEO_CACHE[cache_key]

    out = _geo_no_match(provider)

    # Lazy import (only needed if you actually use nominatim)
    try:
        from geopy.geocoders import Nominatim
        from geopy.extra.rate_limiter import RateLimiter
        from geopy.exc import GeocoderInsufficientPrivileges

        geolocator = Nominatim(
            user_agent="diane-rocher-github-sourcer/1.0 (contact: dianemichaela88@gmail.com)",
            timeout=10
        )

        geocode = RateLimiter(
            geolocator.geocode,
            min_delay_seconds=1.2,   # be gentle
            max_retries=2,
            error_wait_seconds=2.0
        )

        try:
            loc = geocode(raw, addressdetails=True)
        except GeocoderInsufficientPrivileges:
            out["owner_geocode_status"] = "OSM_403_BLOCKED"
            _GEO_CACHE[cache_key] = out
            return out

        if not loc:
            _GEO_CACHE[cache_key] = out
            return out

        addr = (loc.raw or {}).get("address") or {}
        city = addr.get("city") or addr.get("town") or addr.get("village") or ""
        region = addr.get("state") or addr.get("region") or ""
        country = addr.get("country") or ""
        country_code = (addr.get("country_code") or "").upper()

        out = {
            "owner_location_norm": loc.address or "",
            "owner_city": city,
            "owner_region": region,
            "owner_country": country,
            "owner_country_code": country_code,
            "owner_lat": str(loc.latitude),
            "owner_lon": str(loc.longitude),
            "owner_geocode_provider": provider,
            "owner_geocode_status": "OK",
        }

    except Exception:
        out["owner_geocode_status"] = "ERROR"

    _GEO_CACHE[cache_key] = out
    return out


def geocode_and_normalize(raw: str) -> dict:
    """
    Choose provider:
    - If GEO_PROVIDER explicitly set: use it
    - else: use google if API key present; otherwise nominatim
    """
    provider = GEO_PROVIDER
    if not provider:
        provider = "google" if GOOGLE_MAPS_API_KEY else "nominatim"

    if provider == "google":
        return geocode_google(raw)
    if provider == "nominatim":
        return geocode_nominatim(raw)

    d = _geo_empty(provider)
    d["owner_geocode_status"] = "UNKNOWN_PROVIDER"
    return d


# ---------------- Main ----------------
def main():
    _print_auth_diagnostics()
    _print_rate_limit_snapshot()

    load_geo_cache()

    query = build_query()
    print("Query:", query)
    print("Geocoding provider:", GEO_PROVIDER or ("google" if GOOGLE_MAPS_API_KEY else "nominatim"))
    print("Google API key detected:", "YES" if GOOGLE_MAPS_API_KEY else "NO (will use Nominatim unless GEO_PROVIDER=google)")

    rows: list[dict] = []
    seen = 0

    for repo in search_repositories(query):
        if not created_year_in_range(repo.get("created_at", "")):
            continue

        owner = repo.get("owner") or {}
        owner_login = owner.get("login") or ""
        owner_url = owner.get("html_url") or ""

        ojson = fetch_owner(owner_login) if owner_login else {}
        o = owner_fields(ojson) if ojson else {
            "owner_name": "",
            "owner_email": "",
            "owner_location": "",
            "owner_blog": "",
            "owner_x": "",
            "owner_linkedin": "",
            "owner_extra_links": "",
        }

        geo = geocode_and_normalize(o["owner_location"])

        rows.append({
            # Repo
            "repo_full_name": repo.get("full_name", ""),
            "repo_url": repo.get("html_url", ""),
            "description": repo.get("description", "") or "",
            "language": repo.get("language", "") or "",
            "stars": repo.get("stargazers_count", 0),
            "forks": repo.get("forks_count", 0),
            "open_issues": repo.get("open_issues_count", 0),
            "created_at": repo.get("created_at", ""),
            "updated_at": repo.get("updated_at", ""),
            "pushed_at": repo.get("pushed_at", ""),

            # Owner
            "owner_login": owner_login,
            "owner_url": owner_url,
            "owner_name": o["owner_name"],
            "owner_location": o["owner_location"],  # raw
            "owner_email": o["owner_email"],
            "owner_blog": o["owner_blog"],
            "owner_x": o["owner_x"],
            "owner_linkedin": o["owner_linkedin"],
            "owner_extra_links": o["owner_extra_links"],

            # Geocoded owner location
            **geo,
        })

        seen += 1
        if seen >= MAX_REPOS:
            break

        # Save cache periodically
        if seen % 50 == 0:
            save_geo_cache()
            print(f"Progress: {seen} repos (geocode cache saved)")

    df = pd.DataFrame(rows).fillna("")
    write_excel_with_fallback(df, DEFAULT_XLSX)

    save_geo_cache()
    print(f"Geocode cache saved: {GEO_CACHE_FILE.resolve()}")


if __name__ == "__main__":
    main()
