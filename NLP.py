"""
GitHub repo sourcer for LLM fine-tuning / inference — search-budget aware, seed-first.

- Seed repos first (core API only; no search): vLLM, PEFT, bitsandbytes, Axolotl, TRL, Unsloth, etc.
- Search used sparingly: MAX_SEARCH_REQUESTS_PER_RUN, MAX_PAGES_PER_QUERY=1, keyword tiers + rotation.
- On search exhaustion: stop search phase (no long sleep); continue with contributor aggregation.
- Priorities: (1) seed + light search, (2) contributor discovery, (3) person aggregation, (4) expertise_score.
- Core people fields: person_login, person_url, matched_repo_names, matched_keywords, matched_clusters,
  role, contributions, expertise_score. Contact/social optional.
"""

import os
import re
import json
import time
import random
import uuid
import typing as t
from datetime import datetime, timezone, timedelta
from pathlib import Path
from urllib.parse import quote_plus, urlparse

import requests
import pandas as pd

import gspread
from google.oauth2.service_account import Credentials

SPREADSHEET_ID = "1OVr2EigkJ5ZHceilXGn-Zl8xpGTGPVxp8jKIEJnOgmo"
SHEET_GID = 1382489855
SERVICE_ACCOUNT_FILE = "google_service_account.json"

def _gspread_client():
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=scopes
    )
    return gspread.authorize(creds)


def get_gspread_worksheet():
    client = _gspread_client()
    sh = client.open_by_key(SPREADSHEET_ID)
    for ws in sh.worksheets():
        if ws.id == SHEET_GID:
            return ws
    raise ValueError(f"Worksheet with gid={SHEET_GID} not found. Check the gid in your URL.")


def get_or_create_people_worksheet():
    """Get or create worksheet named 'people' in the same spreadsheet."""
    client = _gspread_client()
    sh = client.open_by_key(SPREADSHEET_ID)
    for ws in sh.worksheets():
        if ws.title == "people":
            return ws
    return sh.add_worksheet(title="people", rows=1000, cols=30)


from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from requests.exceptions import ReadTimeout, ConnectionError, HTTPError

from dotenv import load_dotenv
load_dotenv()

import os
print("TOKEN present:", bool(os.getenv("GITHUB_TOKEN")))



# ---------------- Config ----------------
GITHUB_API = "https://api.github.com"
SCRIPT_SOURCE = "NLP"  # written into the 'source' column so multiple scripts can be traced

# --------------- Search budget (GitHub Search API ~30 req/hour authenticated) ---------------
MAX_SEARCH_REQUESTS_PER_RUN = int(os.getenv("MAX_SEARCH_REQUESTS_PER_RUN", "20"))   # stop before exhausting quota
MAX_PAGES_PER_QUERY = int(os.getenv("MAX_PAGES_PER_QUERY", "1"))                    # 1 = one request per keyword
MAX_KEYWORDS_PER_RUN = int(os.getenv("MAX_KEYWORDS_PER_RUN", "10"))                 # cap keywords per run (rotation)
SEARCH_QUOTA_STOP_THRESHOLD = 2   # stop search phase when remaining <= this (do not long-sleep)
ENABLE_CREATED_SCAN = os.getenv("ENABLE_CREATED_SCAN", "false").strip().lower() in ("1", "true", "yes")
ENABLE_PUSHED_SCAN = os.getenv("ENABLE_PUSHED_SCAN", "true").strip().lower() in ("1", "true", "yes")

# Keyword priority tiers (do not run all every time; use top tiers + rotation)
KEYWORDS_HIGH: list[str] = [
    "QLoRA", "LoRA", "PEFT", "Axolotl", "TRL", "Unsloth", "vLLM", "bitsandbytes", "AWQ", "GPTQ",
]
KEYWORDS_MEDIUM: list[str] = [
    "DeepSpeed", "FSDP", "flash-attention", "FlashAttention",
]
KEYWORDS_OPTIONAL: list[str] = [
    "transformers", "huggingface", "fine-tuning", "finetuning", "datasets", "adapter", "adapters",
    "Hugging Face Transformers", "instruction tuning", "SFT", "RLHF", "DPO", "preference optimization",
    "reward model", "TGI", "Text Generation Inference", "TensorRT-LLM", "Triton Inference Server",
    "ONNX Runtime", "quantization", "4-bit", "int8", "low-rank adaptation", "accelerate",
]
# Default: high + medium only (no optional/broad to save budget and reduce noise)
KEYWORDS_DEFAULT_TIERS = ("high", "medium")  # ("high", "medium", "optional") to include optional

# Seed repos: high-signal; use core API only (no search). Always processed first.
SEED_REPOS: list[str] = [
    "vllm-project/vllm",
    "huggingface/peft",
    "bitsandbytes-foundation/bitsandbytes",
    "axolotl-ai-cloud/axolotl",
    "huggingface/trl",
    "unslothai/unsloth",
]

# Cluster name for search results (single cluster for simplicity)
SEARCH_CLUSTER_NAME = "finetuning_inference"
EXCLUDE_TERMS_BY_CLUSTER: dict[str, list[str]] = {"default": []}
LANG_FILTERS: list[str] = [""]

STATE_FILE = Path(__file__).with_name("state.json")
FIRST_RUN_LOOKBACK_DAYS = 183

PER_PAGE = 100
MAX_REPOS = 2000
MAX_REPOS_PER_QUERY = 50
TIMEOUT = 20

DEFAULT_XLSX = "github_repos_llm_engineering_with_owner_details.xlsx"

# Gentle pacing to reduce abuse detection on /search endpoints
PAGE_SLEEP_RANGE = (0.2, 0.8)  # seconds (randomized)

# Contributors (only for new repos; rate-limit safe)
INCLUDE_CONTRIBUTORS = os.getenv("INCLUDE_CONTRIBUTORS", "true").strip().lower() in ("1", "true", "yes")
TOP_N_CONTRIBUTORS = int(os.getenv("TOP_N_CONTRIBUTORS", "5"))
MIN_STARS_FOR_CONTRIB = int(os.getenv("MIN_STARS_FOR_CONTRIB", "0"))  # fetch contributors from all repos for max discovery
REFRESH_CONTRIBUTORS_ON_UPDATE = True  # always refresh for people sheet; contributors fetched for high-star repos
_CONTRIB_CACHE: dict[str, dict] = {}

# Geocoding
GEO_PROVIDER = (os.getenv("GEO_PROVIDER") or "").strip().lower()  # "google" or "nominatim" (optional)
GOOGLE_MAPS_API_KEY = (os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()

# Cache file for geocoding results (disk)
GEO_CACHE_FILE = Path(__file__).with_name("geocode_cache.json")
_GEO_CACHE: dict[str, dict] = {}

# Common non-geocodable locations
_BAD_LOCATIONS = {
    "", "remote", "worldwide", "earth", "somewhere", "internet", "everywhere", "global", "online",
    "anywhere", "planet earth", "the internet", "github", "home",
}

# ---------------- Token -----------------
# Prefer env var GITHUB_TOKEN. Optionally load from token_1.py if you keep secrets locally.
try:
    from token_1 import GITHUB_TOKEN_2
except Exception:
    GITHUB_TOKEN_2 = ""

ENV_TOKEN = (os.getenv("GITHUB_TOKEN") or "").strip()
FILE_TOKEN = (GITHUB_TOKEN_2 or "").strip()
TOKEN = ENV_TOKEN or FILE_TOKEN

SESSION = requests.Session()
SESSION.headers.update({
    "Accept": "application/vnd.github+json",
    "User-Agent": "github-sourcer-script/5.1",
})

# Use Bearer (most robust across token types)
if TOKEN:
    SESSION.headers.update({"Authorization": f"Bearer {TOKEN}"})
else:
    print("No GitHub token found (GITHUB_TOKEN env var or token_1.py:GITHUB_TOKEN_2). "
          "Running unauthenticated => very low rate limit.")

# 🔍 DEBUG (DOIT être hors du if/else)
auth = SESSION.headers.get("Authorization", "")
print("TOKEN present:", bool(TOKEN))
print("Auth scheme:", auth.split(" ")[0] if auth else "NONE")
print("Auth length:", len(auth))

r = SESSION.get("https://api.github.com/rate_limit", timeout=20)
print("rate_limit status:", r.status_code)
print("rate_limit body:", r.text[:200])


# Owner enrichment (cached in-memory)
_OWNER_CACHE: dict[str, dict] = {}
# Negative cache: logins that returned 404 (do not retry in same run)
_OWNER_404_CACHE: set[str] = set()
# Full repo details (topics) cached
_REPO_CACHE: dict[str, dict] = {}


# ---------------- Search budget & rate limit ----------------
class SearchRateLimitExhausted(Exception):
    """Raised when search API quota is exhausted; do not long-sleep, stop search phase."""
    pass


class GitHubNotFound(Exception):
    """Raised when GET returns 404/422; do not retry."""
    pass


def get_rate_limit_snapshot() -> dict:
    """Return {core: {remaining, reset}, search: {remaining, reset}}. Uses cache if recent."""
    try:
        r = SESSION.get(f"{GITHUB_API}/rate_limit", timeout=TIMEOUT)
        if r.status_code != 200:
            return {"core": {"remaining": 0, "reset": 0}, "search": {"remaining": 0, "reset": 0}}
        data = r.json() or {}
        resources = data.get("resources") or {}
        core = resources.get("core") or {}
        search = resources.get("search") or {}
        return {
            "core": {"remaining": core.get("remaining", 0), "reset": core.get("reset", 0)},
            "search": {"remaining": search.get("remaining", 0), "reset": search.get("reset", 0)},
        }
    except Exception:
        return {"core": {"remaining": 0, "reset": 0}, "search": {"remaining": 0, "reset": 0}}


def can_afford_search_request(threshold: int = SEARCH_QUOTA_STOP_THRESHOLD) -> bool:
    """True if search remaining > threshold. Use before each search request."""
    snap = get_rate_limit_snapshot()
    return (snap.get("search") or {}).get("remaining", 0) > threshold


def get_keywords_for_run() -> tuple[list[str], int]:
    """Return (keywords to run this time, next_start_index for state). Uses tiers + rotation."""
    tiers = KEYWORDS_DEFAULT_TIERS
    all_kw: list[str] = []
    if "high" in tiers:
        all_kw.extend(KEYWORDS_HIGH)
    if "medium" in tiers:
        all_kw.extend(KEYWORDS_MEDIUM)
    if "optional" in tiers:
        all_kw.extend(KEYWORDS_OPTIONAL)
    state = load_state()
    start = state.get("keyword_cursor", 0)
    # Cap how many keywords we run this run
    slice_end = min(start + MAX_KEYWORDS_PER_RUN, len(all_kw))
    keywords_this_run = all_kw[start:slice_end]
    next_cursor = slice_end if slice_end < len(all_kw) else 0  # wrap for next run
    return keywords_this_run, next_cursor


# ---------------- Diagnostics ----------------
def _print_auth_diagnostics() -> None:
    source = "env:GITHUB_TOKEN" if ENV_TOKEN else ("token_1.py:GITHUB_TOKEN_2" if FILE_TOKEN else "none")
    print(f"Token source: {source}")
    print(f"Authorization header present: {'Authorization' in SESSION.headers}")


def _print_rate_limit_snapshot() -> None:
    snap = get_rate_limit_snapshot()
    print(f"RateLimit core: remaining={snap['core'].get('remaining')} reset={snap['core'].get('reset')}")
    print(f"RateLimit search: remaining={snap['search'].get('remaining')} reset={snap['search'].get('reset')}")


# ---------------- Helpers ----------------
def safe_output_path(filename: str) -> Path:
    p = Path(__file__).with_name(filename)
    try:
        p.touch(exist_ok=True)
        return p
    except PermissionError:
        return p.with_name(p.stem + "_new" + p.suffix)


def normalize_url(url: str) -> str:
    if not url:
        return ""
    url = str(url).strip()
    if not url:
        return ""
    if not url.lower().startswith(("http://", "https://")):
        url = "https://" + url
    try:
        return url if urlparse(url).netloc else ""
    except Exception:
        return ""


def urls_from_text(text: str) -> list[str]:
    if not text:
        return []
    urls = re.findall(r"(https?://[^\s)]+)", text, flags=re.IGNORECASE)
    out: list[str] = []
    for u in urls:
        nu = normalize_url(u.rstrip(".,);]}>\"'"))
        if nu and nu not in out:
            out.append(nu)
    return out


def extract_first_linkedin(*fields: str) -> str:
    for field in fields:
        if not field:
            continue
        urls = urls_from_text(field)
        for u in urls:
            if "linkedin.com" in u.lower():
                return u
        s = field.strip()
        if "linkedin.com" in s.lower():
            idx = s.lower().find("linkedin.com")
            candidate = s[idx:].split()[0].strip().rstrip(".,);]}>\"'")
            return normalize_url(candidate)
    return ""


def write_excel_with_fallback(df: pd.DataFrame, filename: str) -> Path:
    out = safe_output_path(filename)
    try:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

        wb = load_workbook(out)
        ws = wb.active
        headers = {c.value: i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}

        link_cols = {
            "repo_url": "Repo",
            "owner_url": "Owner",
            "owner_blog": "Website",
            "owner_x": "X",
            "owner_linkedin": "LinkedIn",
        }

        for r in range(2, ws.max_row + 1):
            for col, label in link_cols.items():
                idx = headers.get(col)
                if not idx:
                    continue
                cell = ws.cell(row=r, column=idx)
                val = (cell.value or "").strip()
                if not val:
                    continue
                url = val.split(";")[0].strip()
                if url.lower().startswith(("http://", "https://")):
                    cell.hyperlink = url
                    cell.value = label
                    cell.font = Font(color="0563C1", underline="single")

        wb.save(out)
        print(f"Excel written: {out.resolve()} ({len(df)} rows)")
        return out

    except Exception as e:
        print("Excel failed, falling back to CSV:", e)
        out_csv = out.with_suffix(".csv")
        df.to_csv(out_csv, index=False)
        print(f"CSV written: {out_csv.resolve()} ({len(df)} rows)")
        return out_csv


def get(url: str, no_sleep_on_search_limit: bool = False) -> requests.Response:
    """Resilient GET with retries. If no_sleep_on_search_limit and search 403, raise SearchRateLimitExhausted."""
    max_attempts = 6
    base_sleep = 2.0
    last_exc: Exception | None = None
    is_search = "/search/" in url

    for attempt in range(1, max_attempts + 1):
        try:
            resp = SESSION.get(url, timeout=TIMEOUT)

            if resp.status_code == 401:
                resp.raise_for_status()

            if resp.status_code in (403, 429):
                remaining = resp.headers.get("X-RateLimit-Remaining")
                if is_search and no_sleep_on_search_limit and (remaining == "0" or remaining is None):
                    print("Search API rate limit exhausted. Stopping search phase (no long sleep).")
                    raise SearchRateLimitExhausted("Search quota exhausted")
                reset = resp.headers.get("X-RateLimit-Reset")
                if reset and (remaining == "0" or remaining is None):
                    wait = max(0, int(reset) - int(time.time())) + 2
                    print(f"Rate limit hit. Sleeping {wait}s...")
                    time.sleep(wait)
                    continue

            # 404/422: do not retry (user/repo not found)
            if resp.status_code in (404, 422):
                raise GitHubNotFound(f"HTTP {resp.status_code}: {url}")

            if resp.status_code in (502, 503, 504):
                sleep = base_sleep * (2 ** (attempt - 1)) + random.uniform(0, 1)
                print(f"Transient {resp.status_code}. Retry {attempt}/{max_attempts} in {sleep:.1f}s")
                time.sleep(sleep)
                continue

            resp.raise_for_status()
            return resp

        except GitHubNotFound:
            raise
        except SearchRateLimitExhausted:
            raise
        except (ReadTimeout, ConnectionError, HTTPError) as e:
            last_exc = e
            sleep = base_sleep * (2 ** (attempt - 1)) + random.uniform(0, 1)
            print(f"Error {e}. Retry {attempt}/{max_attempts} in {sleep:.1f}s")
            time.sleep(sleep)

    raise RuntimeError(f"GET failed after retries: {last_exc}")


def keyword_to_query(keyword: str) -> str:
    """Quote only multi-word or hyphenated phrases; leave acronyms unquoted for better recall."""
    k = keyword.strip()
    return f'"{k}"' if (" " in k or "-" in k) else k


def build_cluster_query(
    keyword: str,
    cluster_name: str,
    start: datetime,
    end: datetime,
    time_filter: str,
    lang_filter: str = "",
) -> str:
    """Build GitHub search query: keyword in name/topics/readme/description, date range, per-cluster excludes, optional language."""
    start_date = start.date().isoformat()
    end_date = end.date().isoformat()
    keyword_query = keyword_to_query(keyword)
    base = f"{keyword_query} in:name in:topics in:readme in:description {time_filter}:{start_date}..{end_date} fork:false"
    if lang_filter:
        base = f"{base} {lang_filter}"
    excludes = EXCLUDE_TERMS_BY_CLUSTER.get(cluster_name) or EXCLUDE_TERMS_BY_CLUSTER["default"]
    negatives = " ".join([f'-"{t}"' if " " in t else f'-{t}' for t in excludes])
    return f"{base} {negatives}".strip()


def search_repositories(
    query: str,
    max_pages: int = MAX_PAGES_PER_QUERY,
    search_budget: dict | None = None,
) -> t.Iterable[dict]:
    """Yield repos from search; cap at max_pages (1 request per page). Respect search_budget['used'] and ['max']."""
    page = 1
    while page <= max_pages:
        if search_budget is not None:
            if search_budget.get("used", 0) >= search_budget.get("max", 0):
                break
        if not can_afford_search_request():
            break
        qp = quote_plus(query)
        url = f"{GITHUB_API}/search/repositories?q={qp}&per_page={PER_PAGE}&page={page}"
        try:
            data = get(url, no_sleep_on_search_limit=True).json()
        except SearchRateLimitExhausted:
            break
        if search_budget is not None:
            search_budget["used"] = search_budget.get("used", 0) + 1
        items = data.get("items") or []
        if not items:
            break
        for repo in items:
            yield repo
        page += 1
        time.sleep(random.uniform(*PAGE_SLEEP_RANGE))


def is_probable_bot(login: str, user_type: str | None = None, html_url: str | None = None) -> bool:
    """Skip or mark before enrichment. [bot], 'bot' in name, or type Bot."""
    if not login:
        return False
    login_lower = login.lower()
    if login_lower.endswith("[bot]") or "[bot]" in login_lower:
        return True
    if "bot" in login_lower:
        return True
    if user_type and str(user_type).lower() == "bot":
        return True
    return False


def is_unresolvable_special_account(login: str) -> bool:
    """Known placeholder/special accounts that often 404 or are not real users."""
    if not login:
        return False
    special = {"copilot", "github", "ghost", "dependabot", "renovate"}
    return login.strip().lower() in special


def _minimal_owner_not_found(login: str) -> dict:
    """Cached minimal record for 404/skip so we never retry."""
    return {"_status": "NOT_FOUND", "login": login}


def fetch_owner(login: str) -> dict:
    """Fetch user profile. 404/special/bot: no retry, negative cache, return minimal with _status NOT_FOUND."""
    if not login:
        return {}
    login = login.strip()
    if login in _OWNER_CACHE:
        return _OWNER_CACHE[login]
    if login in _OWNER_404_CACHE:
        return _minimal_owner_not_found(login)
    if is_probable_bot(login) or is_unresolvable_special_account(login):
        _OWNER_404_CACHE.add(login)
        _OWNER_CACHE[login] = _minimal_owner_not_found(login)
        return _OWNER_CACHE[login]
    try:
        data = get(f"{GITHUB_API}/users/{login}").json()
        if not isinstance(data, dict):
            _OWNER_404_CACHE.add(login)
            _OWNER_CACHE[login] = _minimal_owner_not_found(login)
            print(f"Skipping unresolved GitHub user: {login} (no data)")
            return _OWNER_CACHE[login]
        data["_status"] = "OK"
        _OWNER_CACHE[login] = data
        return _OWNER_CACHE[login]
    except GitHubNotFound:
        _OWNER_404_CACHE.add(login)
        _OWNER_CACHE[login] = _minimal_owner_not_found(login)
        print(f"Skipping unresolved GitHub user: {login} (404)")
        return _OWNER_CACHE[login]


def fetch_top_contributors(owner_login: str, repo_name: str, top_n: int) -> dict:
    """GET repos/{owner}/{repo}/contributors. Returns contributors_top, contributors_top_n, contributors_list (for people sheet)."""
    cache_key = f"{owner_login}/{repo_name}:{top_n}"
    if cache_key in _CONTRIB_CACHE:
        return _CONTRIB_CACHE[cache_key]
    url = f"{GITHUB_API}/repos/{owner_login}/{repo_name}/contributors?per_page={top_n}&anon=false"
    try:
        data = get(url).json()
        if not isinstance(data, list):
            result = {"contributors_top": "", "contributors_top_n": 0, "contributors_list": []}
        else:
            logins = [str(c.get("login", "")) for c in data if c.get("login")]
            contrib_list = [{"login": c.get("login", ""), "contributions": c.get("contributions", 0)} for c in data if c.get("login")]
            result = {
                "contributors_top": "; ".join(logins),
                "contributors_top_n": len(logins),
                "contributors_list": contrib_list,
            }
    except GitHubNotFound:
        result = {"contributors_top": "", "contributors_top_n": 0, "contributors_list": []}
    except Exception:
        result = {"contributors_top": "", "contributors_top_n": 0, "contributors_list": []}
    _CONTRIB_CACHE[cache_key] = result
    return result


def fetch_repo_details(owner_login: str, repo_name: str) -> dict:
    """GET repos/{owner}/{repo} for full details (e.g. topics). Cached. 404 -> empty dict."""
    cache_key = f"{owner_login}/{repo_name}"
    if cache_key in _REPO_CACHE:
        return _REPO_CACHE[cache_key]
    try:
        data = get(f"{GITHUB_API}/repos/{owner_login}/{repo_name}").json()
        _REPO_CACHE[cache_key] = data if isinstance(data, dict) else {}
    except GitHubNotFound:
        _REPO_CACHE[cache_key] = {}
    except Exception:
        _REPO_CACHE[cache_key] = {}
    return _REPO_CACHE[cache_key]


def fetch_repo_by_full_name(full_name: str) -> dict | None:
    """Fetch a single repo by full_name (owner/repo) via core API. Returns repo dict like search items, or None if 404."""
    full_name = (full_name or "").strip()
    if "/" not in full_name:
        return None
    parts = full_name.split("/", 1)
    owner_login, repo_name = parts[0], parts[1]
    if not owner_login or not repo_name:
        return None
    try:
        data = get(f"{GITHUB_API}/repos/{owner_login}/{repo_name}").json()
        if not isinstance(data, dict) or not data.get("full_name"):
            return None
        return data
    except GitHubNotFound:
        return None
    except Exception:
        return None


def owner_fields(owner_json: dict) -> dict:
    name = (owner_json.get("name") or "").strip()
    email = (owner_json.get("email") or "").strip()
    location = (owner_json.get("location") or "").strip()
    blog_raw = (owner_json.get("blog") or "").strip()
    blog = normalize_url(blog_raw)
    bio = (owner_json.get("bio") or "").strip()
    twitter_user = (owner_json.get("twitter_username") or "").strip()
    x_url = normalize_url(f"https://twitter.com/{twitter_user}") if twitter_user else ""
    company = (owner_json.get("company") or "").strip()
    location_raw = (owner_json.get("location") or "").strip()
    linkedin = extract_first_linkedin(blog_raw, bio, company, location_raw)
    extra = urls_from_text(bio)
    for known in [blog, linkedin, x_url]:
        if known and known in extra:
            extra.remove(known)
    extra_links = "; ".join(extra)

    return {
        "owner_name": name,
        "owner_email": email,
        "owner_location": location,  # raw
        "owner_blog": blog,
        "owner_x": x_url,
        "owner_linkedin": linkedin,
        "owner_extra_links": extra_links,
    }


# ---------------- Geocoding (cache + providers) ----------------
def load_geo_cache() -> None:
    global _GEO_CACHE
    if GEO_CACHE_FILE.exists():
        try:
            _GEO_CACHE = json.loads(GEO_CACHE_FILE.read_text(encoding="utf-8"))
            if not isinstance(_GEO_CACHE, dict):
                _GEO_CACHE = {}
        except Exception:
            _GEO_CACHE = {}


def save_geo_cache() -> None:
    try:
        GEO_CACHE_FILE.write_text(
            json.dumps(_GEO_CACHE, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    except Exception:
        pass


def _geo_empty(provider: str = "") -> dict:
    return {
        "owner_location_norm": "",
        "owner_city": "",
        "owner_region": "",
        "owner_country": "",
        "owner_country_code": "",
        "owner_lat": "",
        "owner_lon": "",
        "owner_geocode_provider": provider,
        "owner_geocode_status": "EMPTY",
    }


def _geo_no_match(provider: str) -> dict:
    d = _geo_empty(provider)
    d["owner_geocode_status"] = "NO_MATCH"
    return d


def geocode_google(raw: str) -> dict:
    provider = "google"
    raw = (raw or "").strip()
    key = raw.lower()

    if not raw or key in _BAD_LOCATIONS:
        d = _geo_empty(provider)
        d["owner_geocode_status"] = "SKIPPED"
        return d

    if not GOOGLE_MAPS_API_KEY:
        d = _geo_empty(provider)
        d["owner_geocode_status"] = "NO_API_KEY"
        return d

    cache_key = f"{provider}:{key}"
    if cache_key in _GEO_CACHE:
        return _GEO_CACHE[cache_key]

    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": raw, "key": GOOGLE_MAPS_API_KEY}

    out = _geo_no_match(provider)
    try:
        r = requests.get(url, params=params, timeout=TIMEOUT)
        data = r.json() if r.content else {}
        status = (data.get("status") or "").upper()

        if status != "OK":
            out["owner_geocode_status"] = status or "ERROR"
            _GEO_CACHE[cache_key] = out
            return out

        results = data.get("results") or []
        if not results:
            _GEO_CACHE[cache_key] = out
            return out

        top = results[0]
        formatted = top.get("formatted_address") or ""
        geom = (top.get("geometry") or {}).get("location") or {}
        lat = geom.get("lat")
        lon = geom.get("lng")

        comps = top.get("address_components") or []
        comp_map = {}
        for c in comps:
            types = c.get("types") or []
            for ty in types:
                comp_map.setdefault(ty, c)

        def _long(ty: str) -> str:
            return (comp_map.get(ty) or {}).get("long_name") or ""

        def _short(ty: str) -> str:
            return (comp_map.get(ty) or {}).get("short_name") or ""

        city = _long("locality") or _long("postal_town") or _long("administrative_area_level_3")
        region = _long("administrative_area_level_1")
        country = _long("country")
        country_code = _short("country")

        out = {
            "owner_location_norm": formatted,
            "owner_city": city,
            "owner_region": region,
            "owner_country": country,
            "owner_country_code": country_code,
            "owner_lat": "" if lat is None else str(lat),
            "owner_lon": "" if lon is None else str(lon),
            "owner_geocode_provider": provider,
            "owner_geocode_status": "OK",
        }

    except Exception:
        out["owner_geocode_status"] = "ERROR"

    _GEO_CACHE[cache_key] = out
    return out


def geocode_nominatim(raw: str) -> dict:
    provider = "nominatim"
    raw = (raw or "").strip()
    key = raw.lower()

    if not raw or key in _BAD_LOCATIONS:
        d = _geo_empty(provider)
        d["owner_geocode_status"] = "SKIPPED"
        return d

    cache_key = f"{provider}:{key}"
    if cache_key in _GEO_CACHE:
        return _GEO_CACHE[cache_key]

    out = _geo_no_match(provider)

    # Lazy import (only needed if you actually use nominatim)
    try:
        from geopy.geocoders import Nominatim
        from geopy.extra.rate_limiter import RateLimiter
        from geopy.exc import GeocoderInsufficientPrivileges

        geolocator = Nominatim(
            user_agent="diane-rocher-github-sourcer/1.0 (contact: dianemichaela88@gmail.com)",
            timeout=10
        )

        geocode = RateLimiter(
            geolocator.geocode,
            min_delay_seconds=1.2,   # be gentle
            max_retries=2,
            error_wait_seconds=2.0
        )

        try:
            loc = geocode(raw, addressdetails=True)
        except GeocoderInsufficientPrivileges:
            out["owner_geocode_status"] = "OSM_403_BLOCKED"
            _GEO_CACHE[cache_key] = out
            return out

        if not loc:
            _GEO_CACHE[cache_key] = out
            return out

        addr = (loc.raw or {}).get("address") or {}
        city = addr.get("city") or addr.get("town") or addr.get("village") or ""
        region = addr.get("state") or addr.get("region") or ""
        country = addr.get("country") or ""
        country_code = (addr.get("country_code") or "").upper()

        out = {
            "owner_location_norm": loc.address or "",
            "owner_city": city,
            "owner_region": region,
            "owner_country": country,
            "owner_country_code": country_code,
            "owner_lat": str(loc.latitude),
            "owner_lon": str(loc.longitude),
            "owner_geocode_provider": provider,
            "owner_geocode_status": "OK",
        }

    except Exception:
        out["owner_geocode_status"] = "ERROR"

    _GEO_CACHE[cache_key] = out
    return out


def geocode_and_normalize(raw: str) -> dict:
    """
    Choose provider:
    - If GEO_PROVIDER explicitly set: use it
    - else: use google if API key present; otherwise nominatim
    """
    provider = GEO_PROVIDER
    if not provider:
        provider = "google" if GOOGLE_MAPS_API_KEY else "nominatim"

    if provider == "google":
        return geocode_google(raw)
    if provider == "nominatim":
        return geocode_nominatim(raw)

    d = _geo_empty(provider)
    d["owner_geocode_status"] = "UNKNOWN_PROVIDER"
    return d


def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            data = json.loads(STATE_FILE.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def save_state(state: dict) -> None:
    try:
        STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def compute_window() -> tuple[datetime, datetime]:
    """Fixed rolling window: always the last FIRST_RUN_LOOKBACK_DAYS (e.g. 6 months). Ignores state.json."""
    end = datetime.now(timezone.utc).replace(microsecond=0)
    start = end - timedelta(days=FIRST_RUN_LOOKBACK_DAYS)
    return start.replace(microsecond=0), end


def ensure_header(ws, header: list[str]) -> list[str]:
    existing = ws.row_values(1)
    if not existing:
        ws.append_row(header)
        return header

    s = set(existing)
    missing = [c for c in header if c not in s]
    if missing:
        new_header = existing + missing
        ws.update(values=[new_header], range_name="1:1")
        return new_header
    return existing


def load_existing_repo_full_names(ws) -> set[str]:
    header = ws.row_values(1)
    if "repo_full_name" not in header:
        return set()
    col_index = header.index("repo_full_name") + 1
    vals = ws.col_values(col_index)[1:]
    return set(v.strip() for v in vals if v and v.strip())


def append_rows(ws, rows: list[dict], header: list[str]) -> None:
    """Write new rows starting at column A of the next empty row. No append_rows(); explicit A anchor."""
    if not rows:
        return
    next_row = len(ws.col_values(1)) + 1
    values_matrix = [[row.get(col, "") for col in header] for row in rows]
    ws.update(values=values_matrix, range_name=f"A{next_row}", value_input_option="RAW")


def build_repo_row_map(ws) -> dict[str, int]:
    """Return mapping repo_full_name -> 1-based row number (header is row 1)."""
    header = ws.row_values(1)
    if "repo_full_name" not in header:
        return {}
    col_index = header.index("repo_full_name") + 1
    vals = ws.col_values(col_index)[1:]
    out: dict[str, int] = {}
    for i, v in enumerate(vals):
        if v and v.strip():
            out[v.strip()] = i + 2
    return out


CONTRIBUTOR_COLUMNS = ("contributors_top", "contributors_top_n")


def upsert_rows(ws, rows: list[dict], header: list[str]) -> tuple[int, int]:
    """Update existing rows by repo_full_name, append new ones. Returns (num_appended, num_updated)."""
    repo_row = build_repo_row_map(ws)
    num_updated = 0
    to_append: list[dict] = []
    last_col = get_column_letter(len(header))
    updates: list[tuple[int, list]] = []
    for r in rows:
        key = (r.get("repo_full_name") or "").strip()
        if not key:
            continue
        values = [r.get(col, "") for col in header]
        if key in repo_row:
            row_num = repo_row[key]
            if not REFRESH_CONTRIBUTORS_ON_UPDATE:
                try:
                    existing = ws.row_values(row_num)
                    for col in CONTRIBUTOR_COLUMNS:
                        if col in header:
                            idx = header.index(col)
                            if idx < len(existing):
                                values[idx] = existing[idx]
                except Exception:
                    pass
            updates.append((row_num, values))
            num_updated += 1
        else:
            to_append.append(r)
    if updates:
        body = [{"range": f"A{row_num}:{last_col}{row_num}", "values": [values]} for row_num, values in updates]
        ws.batch_update(body, value_input_option="RAW")
    if to_append:
        append_rows(ws, to_append, header)
    return len(to_append), num_updated


# People sheet: core fields first (ranking); optional enrichment last (no impact on ranking)
PEOPLE_HEADER = [
    "person_login",
    "person_url",
    "matched_repo_names",
    "matched_keywords",
    "matched_clusters",
    "role",
    "contributions",
    "expertise_score",
    "person_profile_status",
    "person_name",
    "person_email",
    "person_location",
    "person_blog",
    "person_x",
    "person_linkedin",
    "run_id",
    "run_timestamp_utc",
]


def _compute_expertise_score(
    n_repos: int,
    n_keywords: int,
    n_clusters: int,
    total_contributions: int,
    is_owner: bool,
) -> float:
    """Score for ranking; higher = stronger candidate. Does not use contact/social data."""
    score = n_repos * 2.0 + n_keywords * 1.0 + n_clusters * 2.0
    score += min(total_contributions, 500) / 10.0  # cap contrib effect
    if is_owner:
        score += 5.0
    return round(score, 1)


def _aggregate_people(
    raw_people: list[dict],
) -> list[dict]:
    """Deduplicate by person_login; aggregate repos, keywords, clusters; compute expertise_score."""
    # raw_people: list of {login, role, contributions, repo, cluster, keyword}
    by_login: dict[str, dict] = {}
    for p in raw_people:
        login = (p.get("login") or "").strip()
        if not login:
            continue
        if login not in by_login:
            by_login[login] = {
                "repos": [],
                "keywords": set(),
                "clusters": set(),
                "roles": set(),
                "total_contributions": 0,
            }
        by_login[login]["repos"].append(p.get("repo", ""))
        by_login[login]["keywords"].add(p.get("keyword", ""))
        by_login[login]["clusters"].add(p.get("cluster", ""))
        by_login[login]["roles"].add(p.get("role", ""))
        by_login[login]["total_contributions"] += p.get("contributions", 0)
    out: list[dict] = []
    for login, ag in by_login.items():
        repos = [r for r in ag["repos"] if r]
        keywords = sorted(ag["keywords"] - {""})
        clusters = sorted(ag["clusters"] - {""})
        roles = sorted(ag["roles"] - {""})
        total_contrib = ag["total_contributions"]
        is_owner = "owner" in ag["roles"]
        expertise_score = _compute_expertise_score(
            len(repos), len(keywords), len(clusters), total_contrib, is_owner
        )
        out.append({
            "login": login,
            "matched_repo_names": "; ".join(repos),
            "matched_keywords": "; ".join(keywords),
            "matched_clusters": "; ".join(clusters),
            "role": "; ".join(roles) if roles else "",
            "contributions": total_contrib,
            "expertise_score": expertise_score,
        })
    return out


def _build_aggregated_person_row(
    ag: dict,
    run_id: str,
    run_ts: str,
    fetch_profile: bool = True,
) -> dict:
    """Build one people-sheet row from aggregated person data. Optional lightweight profile (no geocode)."""
    login = ag.get("login", "")
    if not login:
        return {}
    person_profile_status = "OK"
    person_name = ""
    person_email = ""
    person_location = ""
    person_blog = ""
    person_x = ""
    person_linkedin = ""
    if fetch_profile:
        ojson = fetch_owner(login)
        person_profile_status = ojson.get("_status", "OK") if ojson else "NOT_FOUND"
        o = owner_fields(ojson) if ojson else {}
        person_name = o.get("owner_name", "")
        person_email = o.get("owner_email", "")
        person_location = o.get("owner_location", "")
        person_blog = o.get("owner_blog", "")
        person_x = o.get("owner_x", "")
        person_linkedin = o.get("owner_linkedin", "")
    return {
        "person_login": login,
        "person_url": f"https://github.com/{login}" if login else "",
        "matched_repo_names": ag.get("matched_repo_names", ""),
        "matched_keywords": ag.get("matched_keywords", ""),
        "matched_clusters": ag.get("matched_clusters", ""),
        "role": ag.get("role", ""),
        "contributions": ag.get("contributions", 0),
        "expertise_score": ag.get("expertise_score", 0),
        "person_profile_status": person_profile_status,
        "person_name": person_name,
        "person_email": person_email,
        "person_location": person_location,
        "person_blog": person_blog,
        "person_x": person_x,
        "person_linkedin": person_linkedin,
        "run_id": run_id,
        "run_timestamp_utc": run_ts,
    }


def _build_row_from_repo(
    repo: dict,
    skill_cluster: str,
    keyword_matched: str,
    full_query: str,
    scan_type: str,
) -> dict:
    """Build one sheet row from a GitHub repo with traceability and repo_topics."""
    owner = repo.get("owner") or {}
    owner_login = owner.get("login") or ""
    owner_url = owner.get("html_url") or ""
    full_name = repo.get("full_name", "")
    repo_name = repo.get("name") or (full_name.split("/")[-1] if full_name else "")
    ojson = fetch_owner(owner_login) if owner_login else {}
    o = owner_fields(ojson) if ojson else {
        "owner_name": "", "owner_email": "", "owner_location": "", "owner_blog": "",
        "owner_x": "", "owner_linkedin": "", "owner_extra_links": "",
    }
    geo = geocode_and_normalize(o["owner_location"])
    # Full repo details for topics (cached)
    repo_details = fetch_repo_details(owner_login, repo_name) if owner_login and repo_name else {}
    topics_list = repo_details.get("topics") or []
    repo_topics = "; ".join(topics_list) if isinstance(topics_list, list) else ""
    return {
        "repo_full_name": full_name,
        "repo_url": repo.get("html_url", ""),
        "repo_topics": repo_topics,
        "description": repo.get("description", "") or "",
        "language": repo.get("language", "") or "",
        "stars": repo.get("stargazers_count", 0),
        "forks": repo.get("forks_count", 0),
        "open_issues": repo.get("open_issues_count", 0),
        "created_at": repo.get("created_at", ""),
        "updated_at": repo.get("updated_at", ""),
        "pushed_at": repo.get("pushed_at", ""),
        "owner_login": owner_login,
        "owner_url": owner_url,
        "owner_name": o["owner_name"],
        "owner_location": o["owner_location"],
        "owner_email": o["owner_email"],
        "owner_blog": o["owner_blog"],
        "owner_x": o["owner_x"],
        "owner_linkedin": o["owner_linkedin"],
        "owner_extra_links": o["owner_extra_links"],
        **geo,
        "skill_cluster": skill_cluster,
        "keyword_matched": keyword_matched,
        "query": full_query,
        "scan_type": scan_type,
        "test_run_marker": datetime.now(timezone.utc).isoformat(),
        "contributors_top": "",
        "contributors_top_n": 0,
    }


# ---------------- Main ----------------
def main():
    _print_auth_diagnostics()
    load_geo_cache()

    # Search budget summary at startup
    snap = get_rate_limit_snapshot()
    search_remaining = (snap.get("search") or {}).get("remaining", 0)
    core_remaining = (snap.get("core") or {}).get("remaining", 0)
    safe_queries = min(max(0, search_remaining - SEARCH_QUOTA_STOP_THRESHOLD), MAX_SEARCH_REQUESTS_PER_RUN)
    print("--- Rate limit ---")
    print(f"Search API: remaining={search_remaining} (stop when <={SEARCH_QUOTA_STOP_THRESHOLD})")
    print(f"Core API: remaining={core_remaining}")
    print(f"Max search requests this run: {MAX_SEARCH_REQUESTS_PER_RUN}; estimated safe: {safe_queries}")
    print(f"Max pages per query: {MAX_PAGES_PER_QUERY}; created_scan={ENABLE_CREATED_SCAN}; pushed_scan={ENABLE_PUSHED_SCAN}")

    window_start, window_end = compute_window()
    print("Window:", window_start.isoformat(), "->", window_end.isoformat())

    header = [
        "source",
        "run_id", "run_timestamp_utc", "window_start_utc", "window_end_utc",
        "skill_cluster", "keyword_matched", "query", "scan_type", "test_run_marker",
        "repo_full_name", "repo_url", "repo_topics", "description", "language", "stars", "forks", "open_issues",
        "created_at", "updated_at", "pushed_at",
        "owner_login", "owner_url", "owner_name", "owner_location", "owner_email",
        "owner_blog", "owner_x", "owner_linkedin", "owner_extra_links",
        "contributors_top", "contributors_top_n",
        "owner_location_norm", "owner_city", "owner_region", "owner_country", "owner_country_code",
        "owner_lat", "owner_lon", "owner_geocode_provider", "owner_geocode_status",
    ]
    ws = get_gspread_worksheet()
    print("Writing to worksheet:", ws.title, "gid:", ws.id)
    header = ensure_header(ws, header)
    repo_row = build_repo_row_map(ws)

    all_rows: dict[str, dict] = {}
    search_queries_skipped_budget = 0

    # ---------- 1. Seed repos first (core API only; no search)
    print("--- Seed repos (core API) ---")
    for full_name in SEED_REPOS:
        repo = fetch_repo_by_full_name(full_name)
        if not repo:
            print(f"  Skip (not found or error): {full_name}")
            continue
        row = _build_row_from_repo(repo, "seed", full_name, "seed", "seed")
        key = (row.get("repo_full_name") or "").strip()
        if key:
            all_rows[key] = row
            print(f"  Added: {key}")
    print(f"Seed repos added: {len([k for k in all_rows if all_rows[k].get('skill_cluster') == 'seed'])}")

    # ---------- 2. Search phase (budget-aware; optional created/pushed)
    keywords_this_run, next_cursor = get_keywords_for_run()
    search_budget = {"used": 0, "max": MAX_SEARCH_REQUESTS_PER_RUN}

    def run_scan(scan_type: str) -> None:
        nonlocal search_queries_skipped_budget
        for keyword in keywords_this_run:
            if search_budget["used"] >= search_budget["max"]:
                search_queries_skipped_budget += 1
                continue
            if not can_afford_search_request():
                break
            for lang_filter in LANG_FILTERS:
                if search_budget["used"] >= search_budget["max"]:
                    break
                q = build_cluster_query(keyword, SEARCH_CLUSTER_NAME, window_start, window_end, scan_type, lang_filter=lang_filter)
                lang_label = lang_filter or "(no lang filter)"
                print(f"[{scan_type}] {SEARCH_CLUSTER_NAME} / {keyword!r} ({lang_label}): budget used={search_budget['used']}/{search_budget['max']}")
                seen_this_query = 0
                try:
                    for repo in search_repositories(q, max_pages=MAX_PAGES_PER_QUERY, search_budget=search_budget):
                        row = _build_row_from_repo(repo, SEARCH_CLUSTER_NAME, keyword, q, scan_type)
                        key = (row.get("repo_full_name") or "").strip()
                        if key:
                            all_rows[key] = row
                        seen_this_query += 1
                        if seen_this_query >= MAX_REPOS_PER_QUERY:
                            break
                        if len(all_rows) >= MAX_REPOS:
                            break
                except SearchRateLimitExhausted:
                    print("  Search quota exhausted; stopping search phase.")
                    return
                if seen_this_query > 0:
                    save_geo_cache()

    if ENABLE_PUSHED_SCAN:
        print("--- Search (pushed) ---")
        run_scan("pushed")
    if ENABLE_CREATED_SCAN and search_budget["used"] < search_budget["max"]:
        print("--- Search (created) ---")
        run_scan("created")

    search_requests_used = search_budget["used"]
    print(f"Search phase: requests used={search_requests_used}; queries skipped (budget)={search_queries_skipped_budget}")

    run_id = uuid.uuid4().hex[:10]
    run_ts = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    for r in all_rows.values():
        r["source"] = SCRIPT_SOURCE
        r["run_id"] = run_id
        r["run_timestamp_utc"] = run_ts
        r["window_start_utc"] = window_start.isoformat()
        r["window_end_utc"] = window_end.isoformat()

    # Contributor discovery: fetch for all repos (MIN_STARS_FOR_CONTRIB=0 for max discovery)
    raw_people: list[dict] = []  # {login, role, contributions, repo, cluster, keyword}
    contributors_fetched_count = 0
    for key, r in all_rows.items():
        if not INCLUDE_CONTRIBUTORS:
            continue
        if (r.get("stars") or 0) < MIN_STARS_FOR_CONTRIB:
            continue
        owner_login = r.get("owner_login", "")
        cluster = r.get("skill_cluster", "")
        keyword = r.get("keyword_matched", "")
        parts = key.split("/", 1)
        repo_name = parts[1] if len(parts) == 2 else ""
        if not owner_login or not repo_name:
            continue
        try:
            data = fetch_top_contributors(owner_login, repo_name, TOP_N_CONTRIBUTORS)
            r["contributors_top"] = data.get("contributors_top", "")
            r["contributors_top_n"] = data.get("contributors_top_n", 0)
            contributors_fetched_count += 1
            raw_people.append({"login": owner_login, "role": "owner", "contributions": 0, "repo": key, "cluster": cluster, "keyword": keyword})
            for c in data.get("contributors_list", []):
                raw_people.append({
                    "login": c.get("login", ""),
                    "role": "contributor",
                    "contributions": c.get("contributions", 0),
                    "repo": key,
                    "cluster": cluster,
                    "keyword": keyword,
                })
        except Exception:
            r["contributors_top"] = ""
            r["contributors_top_n"] = 0
    print(f"Contributors: fetched for {contributors_fetched_count} repos; raw person-repo links: {len(raw_people)}")

    # Person aggregation: dedupe by login; aggregate repos/keywords/clusters; compute expertise_score
    aggregated = _aggregate_people(raw_people)
    aggregated.sort(key=lambda x: (x.get("expertise_score", 0), x.get("login", "")), reverse=True)

    num_appended, num_updated = upsert_rows(ws, list(all_rows.values()), header)

    # People sheet: one row per person (ranked by expertise_score); optional profile fields
    ws_people = get_or_create_people_worksheet()
    ensure_header(ws_people, PEOPLE_HEADER)
    people_rows = [_build_aggregated_person_row(ag, run_id, run_ts, fetch_profile=True) for ag in aggregated]
    people_to_append = [row for row in people_rows if row]
    if people_to_append:
        append_rows(ws_people, people_to_append, PEOPLE_HEADER)

    # Summary
    print("---")
    print(f"Unique repos found: {len(all_rows)}")
    print(f"Repos appended to sheet: {num_appended}. Updated: {num_updated}.")
    print(f"Unique people (after aggregation): {len(aggregated)}; rows appended to 'people' sheet: {len(people_to_append)}")

    state = load_state()
    state["last_successful_created_scan_utc"] = window_end.isoformat()
    state["last_successful_pushed_scan_utc"] = window_end.isoformat()
    state["keyword_cursor"] = next_cursor  # for next run's keyword rotation
    save_state(state)
    print("Saved state (keyword_cursor for next run):", next_cursor)

    save_geo_cache()
    print(f"Geocode cache saved: {GEO_CACHE_FILE.resolve()}")

    # Final search-budget summary
    print("---")
    print("Search-budget summary: total search requests used this run:", search_budget["used"])
    snap_end = get_rate_limit_snapshot()
    print("Search API remaining after run:", (snap_end.get("search") or {}).get("remaining", "?"))


if __name__ == "__main__":
    main()